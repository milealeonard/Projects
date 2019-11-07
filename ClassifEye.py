import glob
import os
import string

import wx
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

class Main_Menu (wx.Panel):
    def __init__ (self, parent):
        super().__init__(parent)
        self.initialise_inputs(parent)
        self.layout(parent)
    def initialise_inputs(self, parent):
        "Initialise the inputs for Main_Menu"
        self.parent = parent
        self.photos = []
        self.users = []
        self.current_user = 0
        self.username_text = ""
        self.renaming_text = ""
        self.have_found = False
        self.wb = None #Current Workbook
        self.ws = None #Current Worksheet
        self.wb_path = None #Current Workbook Path
        self.photo_counter = 0
        self.user_counter = 0
        self.list = []
    def layout (self, parent):
        "Create the basic layout with sizers, buttons, and text boxes"
        self.initiate_sizers(parent)
        self.initiate_widgets(parent)
        self.exit_btn = wx.Button (self, label = "Exit")
        self.exit_btn.Bind (wx.EVT_BUTTON, parent.dstroy)
        self.main_sizer.Add(self.exit_btn, 0, wx.CENTER|wx.ALL, 5)
        self.sizer_right.Add (self.sizer_select_file, 0, wx.ALIGN_LEFT|wx.ALL, 15)
        self.sizer_right.Add (self.sizer_select_folder, 0, wx.ALIGN_LEFT|wx.ALL, 15)
        self.sizer_right.Add (self.sizer_two, 0, wx.ALIGN_LEFT|wx.ALL, 15)
        self.sizer_one.Add(self.sizer_left, 1, wx.ALIGN_RIGHT, 15)
        self.sizer_one.Add(self.sizer_right, 1, wx.ALIGN_LEFT, 15)
        self.main_sizer.Add(self.sizer_one, 0, wx.CENTER, 15)
        self.main_sizer.Add (self.next_frame, 0, wx.CENTER|wx.ALL, 15)
        self.SetSizer (self.main_sizer)
        self.create_excel_file_sizer()
        self.select_folder_sizer()
        self.select_user_sizer()
    def initiate_sizers(self, parent):
        self.main_sizer = wx.BoxSizer(wx.VERTICAL)
        self.sizer_one = wx.BoxSizer (wx.HORIZONTAL)
        self.sizer_left = wx.BoxSizer(wx.VERTICAL)
        self.sizer_right = wx.BoxSizer(wx.VERTICAL)
        self.sizer_select_folder = wx.BoxSizer(wx.HORIZONTAL)
        self.sizer_select_file = wx.BoxSizer(wx.HORIZONTAL)
        self.sizer_two = wx.BoxSizer(wx.HORIZONTAL)
    def initiate_widgets(self, parent):
        self.user_slct = wx.TextCtrl (self, size = (245, -1))
        self.user_slct.Bind (wx.EVT_KEY_DOWN, self.user_selection)
        self.user_slct.SetValue("Type Username Here, then Press Enter")
        self.name_box = wx.TextCtrl (self, size = (210, -1))
        self.name_box.Bind (wx.EVT_KEY_DOWN, self.rename_func)
        self.name_box.SetValue("Type File Name, then Press Enter")
        self.name_box.Hide()
        self.next_frame = wx.Button (self, label = "Next")
        self.next_frame.Bind(wx.EVT_BUTTON, parent.one_to_two)
        self.folder_slct = wx.Button (self, label = "Folder")
        self.folder_slct.Bind (wx.EVT_BUTTON, self.folder_selection)
        self.create_button = wx.Button (self, label = "Create New")
        self.create_button.Bind (wx.EVT_BUTTON, self.naming_time)
        self.browse_button = wx.Button (self, label = "Browse")
        self.browse_button.Bind (wx.EVT_BUTTON, self.browse_function)
    def create_excel_file_sizer (self):
        "Create the Sizer for create excel file"
        self.select_excel_static_text = wx.StaticText (self, -1, label = "Step 1: Browse for Existing Excel File, or Create a New One", )
        self.select_excel_static_text.SetForegroundColour((250,148,18)) # set text color
        self.sizer_left.Add (self.select_excel_static_text, 0, wx.ALIGN_RIGHT|wx.ALL, 32)
        self.sizer_select_file.Add (self.browse_button, 0, wx.CENTER|wx.ALL, 15)
        self.sizer_select_file.Add (self.name_box, 0, wx.CENTER|wx.ALL, 15)
        self.sizer_select_file.Add (self.create_button, 0, wx.CENTER|wx.ALL, 15)
    def select_folder_sizer(self):
        self.select_folder_static_text = wx.StaticText (self, -1, "Step 2: Select Image Directory", )
        self.select_folder_static_text.SetForegroundColour((250,148,18)) # set text color
        self.sizer_left.Add (self.select_folder_static_text, 0, wx.ALIGN_RIGHT|wx.ALL, 32)
        self.sizer_select_folder.Add (self.folder_slct, 0, wx.ALIGN_LEFT|wx.ALL, 15)
    def select_user_sizer(self):
        self.select_user_static_text = wx.StaticText (self, -1, "Step 3: Select Current User", )
        self.select_user_static_text.SetForegroundColour((250,148,18)) # set text color
        self.sizer_left.Add (self.select_user_static_text, 0, wx.ALIGN_RIGHT|wx.ALL, 32)
        self.sizer_two.Add (self.user_slct, 0, wx.ALIGN_LEFT|wx.ALL, 15)
    def rename_func (self, event):
        "Name the newly created Excel file"
        keycode = event.GetKeyCode()
        if chr(keycode) in string.ascii_letters or chr(keycode) in string.digits or keycode == 13:
            if keycode == 13:
                self.excel_name = self.renaming_text
                if self.excel_name[-5:] != ".XLSX":
                    self.excel_name += ".XLSX"
                    self.create_function(event)
            else:
                self.renaming_text += chr(keycode)
            self.name_box.SetValue(self.renaming_text)
        elif keycode == 8:
            n = len (self.renaming_text) - 1
            self.renaming_text = self.renaming_text[0:n]
            self.name_box.SetValue(self.renaming_text)
    def naming_time(self, event):
        self.name_box.Show()
        self.name_box.SetFocus()
        self.create_button.Hide()
        self.Layout()
    def create_function (self, event):
        "Create New Excel file with name from rename_func"
        with wx.DirDialog (self, "Choose Where to Save Your New File") as dlg:
            if dlg.ShowModal () == wx.ID_OK:
                self.wb = Workbook()
                self.ws = self.wb.active
                self.wb_path = dlg.GetPath()+"/"+self.excel_name
        if self.wb_path != None:
            self.select_excel_static_text.SetForegroundColour((3,250,0))
            if self.parent.excel_static_only_once is False:
                self.parent.excel_check.Hide()
                self.parent.excel_advice.Hide()
                self.parent.excel_advice_two.Hide()
                self.parent.Layout()
    def browse_function(self, event):
        "Browse for an existing excel file"
        with wx.FileDialog (None, "Select an Excel File", wildcard = "(*.xlsx)|*.xlsx") as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                self.wb_path = dlg.GetPath()
                self.wb = load_workbook(self.wb_path)
                self.ws = self.wb.active
                for k in range (self.ws.max_column-1):
                    self.users.append(self.ws.cell(row= 1, column=k+3).value) #compile the list of users already in excel file
                c = 2
                while self.ws.cell(row = c, column = 1).value is not None:
                    self.list.append (self.ws.cell(row = c, column = 1).value) #get the list of categories from the excel file
                    c+=1
        if self.wb_path != None:
            self.select_excel_static_text.SetForegroundColour((3,250,0))
            if self.parent.excel_static_only_once is False:
                self.parent.excel_check.Hide()
                self.parent.excel_advice.Hide()
                self.parent.excel_advice_two.Hide()
                self.parent.Layout()
    def user_selection (self, event):
        "Type in the name of the current user, and set the edited column to the selected user"
        keycode = event.GetKeyCode()
        if chr(keycode) in string.ascii_letters or chr(keycode) in string.digits or keycode == 13:
            if keycode == 13:
                self.enter_func(event)
            else:
                self.have_found = False
                self.username_text += chr(keycode)
                self.user_slct.SetValue(self.username_text)
                h = len(self.username_text)
                while self.have_found is False and self.user_counter < len (self.users):
                    for user in self.users:
                        if not user: #skip the "None" users
                            pass
                        elif self.username_text == str(user[0:h]): #Autoprediction
                            self.have_found = True
                            self.user_slct.SetValue(self.username_text+str(user[h::]))
                            break
                        self.user_counter += 1
        elif keycode == 8: #backspace
            self.username_text = self.username_text[0:len(self.username_text)-1]
            self.user_slct.SetValue(self.username_text)
    def enter_func (self, event):
        "When enter is pressed, either open the next panel, or add the current user to the list of users"
        if self.current_user != 0 and self.username_text == "":
            self.parent.one_to_two(event)
        elif self.user_slct.GetValue() in self.users:
            self.current_user = self.users.index(self.user_slct.GetValue()) + 1 # set the current_user to current user index
            self.username_text = ""
        else:
            self.users.append (self.user_slct.GetValue())
            self.ws.cell(row= 1, column=self.ws.max_column+1, value=self.user_slct.GetValue()) # Add the user to the excel file
            self.username_text = ""
            self.current_user = len (self.users)
            self.wb.save(self.wb_path)
        self.select_user_static_text.SetForegroundColour((3,250,0))
        if self.parent.user_static_only_once is False:
            self.parent.user_check.Hide()
            self.parent.Layout()
    def folder_selection (self, event):
        "Choose the image directory that you will be looking at"
        if self.select_excel_static_text.GetForegroundColour()!=((3,250,0)):
            if self.parent.excel_static_only_once:
                self.parent.excel_check = wx.StaticText (self, -1, label = "Please Select an Excel File")
                self.parent.excel_check.SetForegroundColour((255,0,0))
                self.parent.excel_advice = wx.StaticText (self, -1, label = "Tip: If creating a new excel file, remember to press enter after typing its name")
                self.parent.excel_advice.SetForegroundColour((246,250,0))
                self.parent.excel_advice_two = wx.StaticText (self, -1, label = "Tip 2: Remember to press 'Open' after finding where you would like to save the file")
                self.parent.excel_advice_two.SetForegroundColour((246,250,0))
                self.main_sizer.Add(self.parent.excel_check,0, wx.CENTER|wx.ALL, 50)
                self.main_sizer.Add(self.parent.excel_advice,0, wx.CENTER|wx.ALL, 0)
                self.main_sizer.Add(self.parent.excel_advice_two,0, wx.CENTER|wx.ALL, 10)
                self.Layout()
                self.parent.excel_static_only_once = False
        else:
            with wx.DirDialog (self, "Choose an Image Directory") as dlg:
                if dlg.ShowModal () == wx.ID_OK:
                    self.folder_path = dlg.GetPath()
                    self.photos = sorted(glob.glob(os.path.join(self.folder_path, "*jpg")))
                    self.photos += sorted(glob.glob(os.path.join(self.folder_path, "*tif")))
                    self.photos += sorted(glob.glob(os.path.join(self.folder_path, "*png")))
                    self.photos += sorted(glob.glob(os.path.join(self.folder_path, "*ppm")))
                    while self.photo_counter < len (self.photos):
                        for photo in self.photos:
                            photo_name = os.path.basename(photo)
                            self.ws.cell(row= self.photo_counter+2, column=2, value=photo_name) #add the photo names to the excel file
                            self.photo_counter += 1
                    self.wb.save(self.wb_path)
                    if not self.photos:
                        print ("No Photos")
            if self.photos != []:
                self.select_folder_static_text.SetForegroundColour((3,250,0))
            if self.parent.folder_static_only_once is False:
                print ("Yes")
                self.parent.folder_check.Hide()
                self.parent.Layout()
            self.user_slct.SetFocus()
class Categories_Page (wx.Panel):
    def __init__(self, parent, list = []):
        super().__init__(parent)
        self.initialise_inputs(parent, list)
        self.layout(parent)
        self.list_implementation(self.list)
        self.create_new_textbox(self.current_text)
    def initialise_inputs(self, parent, list):
        "Initialise the inputs for Categories_Page"
        self.category_number = len (list)
        self.textbox_count = 0
        self.subcat_number = 0
        self.parent = parent
        self.current_text = str(self.category_number)+": "
        self.subcat_text = str(self.category_number)+"."+str(self.subcat_number)+ ": "
        self.list = list
        self.is_category = True
    def layout(self, parent):
        "Create the basic layout with sizers, buttons, and text boxes"
        self.sizer = wx.BoxSizer (wx.VERTICAL)
        self.sub_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.text_sizer = wx.BoxSizer(wx.VERTICAL)
        self.button_sizer = wx.BoxSizer(wx.VERTICAL)
        self.next_frame_three = wx.Button (self, label = "Next")
        self.next_frame_three.Bind(wx.EVT_BUTTON, parent.two_to_three)
        self.exit_btn = wx.Button (self, label = "Exit")
        self.exit_btn.Bind (wx.EVT_BUTTON, parent.dstroy)
        self.sizer.Add(self.exit_btn, 0, wx.CENTER|wx.ALL, 5)
        self.sizer.Add(self.next_frame_three, 0, wx.CENTER|wx.ALL, 5)
        instructions_panel_two = wx.StaticText(self, label = "Step 4: Define Your Categories")
        instructions_panel_two.SetForegroundColour((250,246,0))
        self.sizer.Add(instructions_panel_two, 0, wx.CENTER|wx.ALL, 5)
        self.sizer.Add(self.sub_sizer, 0, wx.CENTER|wx.ALL, 5)
        self.sub_sizer.Add(self.text_sizer, 1, wx.CENTER|wx.ALL, 5)#Change this to make it non center (1--> 0)
        self.sub_sizer.Add(self.button_sizer, 1, wx.CENTER|wx.ALL, 5) #Change this to make it non center (1--> 0)
        self.SetSizer(self.sizer)
    def list_implementation (self, list):
        "Show and implement the previous list of categories in the textboxes"
        k = 0
        while k < len(list):
            element_name = list[k]
            if "." in element_name:
                self.create_new_textbox(element_name, size = (120, -1))
            else:
                self.create_new_textbox(element_name)
            k+=1
            self.category_number = int(element_name[0]) +1
        self.current_text = str(self.category_number)+": "
    def create_new_textbox(self, value, size = (150,-1), create_plus = 1):
        "create a new textbox with value"
        self.textbox_count +=1
        self.text_box = wx.TextCtrl(self, size = size)
        self.text_box.Bind (wx.EVT_KEY_DOWN, self.typing)
        self.text_box.SetValue(value)
        self.text_sizer.Add(self.text_box, 1,  wx.ALIGN_RIGHT|wx.ALL, 5)
        if create_plus:
            current_sizer = wx.BoxSizer(wx.HORIZONTAL)
            cat_button = wx.Button(self, label = "Add Category", size = (150,-1))
            cat_button.Bind(wx.EVT_BUTTON, self.cat_function)
            current_sizer.Add(cat_button, 1, wx.CENTER|wx.ALL, 5)
            # sub_button = wx.Button(self, label = "Add Subcategory", size = (120,-1))
            # sub_button.Bind(wx.EVT_BUTTON, self.plus_sub)
            # current_sizer.Add(sub_button, 0, wx.CENTER|wx.ALL, 5)
            self.button_sizer.Add(current_sizer, 1, wx.CENTER|wx.ALL, 0)
        else:
            cat_button = wx.Button(self, label = "Add Category", size = (150,-1))
            cat_button.Bind(wx.EVT_BUTTON, self.enter_function)
            self.button_sizer.Add(cat_button, 1, wx.ALIGN_LEFT|wx.ALL, 5)
        self.Layout()
    def cat_function(self, event):
        self.is_category = True
        self.enter_function (event, type = "nonsub")
    # def plus_sub(self, event):
    #     "Create subcategories"
    #     self.is_category = False
    #     self.enter_function (event,type = "subcategory")
    def typing (self, event):
        "Register the key events"
        keycode = event.GetKeyCode()
        if self.is_category:
            if keycode == 13:
                self.enter_function(event, type = "nonsub")
            elif keycode == 8:
                n = len (self.current_text) - 1
                self.current_text = self.current_text[0:n]
            elif chr(keycode) in string.digits or chr(keycode) in string.ascii_letters or keycode == 32:
                self.current_text += chr(keycode)
            self.text_box.SetValue(self.current_text)
        # else:
        #     if keycode == 13:
        #         self.enter_function(event, type = "subcategory")
        #     elif keycode == 8:
        #         n = len (self.subcat_text) - 1
        #         self.subcat_text = self.subcat_text[0:n]
        #     elif chr(keycode) in string.digits or chr(keycode) in string.ascii_letters or keycode == 32:
        #         self.subcat_text += chr(keycode)
        #     self.text_box.SetValue(self.subcat_text)
    def check(self):
        if len(self.list)<self.textbox_count:
            if len(self.text_box.GetValue()) != 3:
                self.list.append(self.text_box.GetValue())
    def enter_function (self, event, type = "nonsub"):
        "When enter is pressed, either go to next panel or add the category to list of categories"
        if type == "nonsub":
            self.subcat_number = 0
            self.list.append (self.current_text)
            self.category_number +=1
            self.current_text = str(self.category_number) + ": "
            self.create_new_textbox(self.current_text)
        # elif type == "subcategory":
        #     if self.subcat_text[0:3] == "0.0":
        #         self.list.append (self.current_text)
        #     else:
        #         self.list.append (self.subcat_text.split(" ")[0:1][0] + " " + self.current_text.split(" ")[1:][0] + " " + self.subcat_text.split(" ")[1:][0])
        #     self.subcat_number +=1
        #     self.subcat_text = str(self.category_number)+"."+str(self.subcat_number)+ ": "
        #     self.create_new_textbox(str(self.category_number)+"."+str(self.subcat_number)+ ": ", size = (100,-1), create_plus = 0)
class Excel_Editor (wx.Panel):
    def __init__ (self, parent, photos = [], excel = None, user = 0, list = []):
        super().__init__(parent)
        self.max_size = 600
        self.current_photo = 0
        self.finding_counter = 0
        self.total_photos = len(photos)
        self.previous_number = None
        self.photos = photos
        self.current_text = "Type Here"
        self.column_no = user+1
        self.list = list
        self.layout(parent)
        self.wb_path = excel
        self.wb = load_workbook(self.wb_path)
        self.ws = self.wb.active
        self.insert_list_into_excel(parent)
        self.wb.save(self.wb_path)
        self.have_found = False
        self.find_row_no(parent)
        if self.photos:
            if self.current_photo>=len(self.photos):
                print ("List Index Out of Range")
            else:
                self.update_photo(self.photos[self.current_photo]) #display the photo corresponding to the first non marked photo
        self.list_implementation(self.list)
    def find_row_no(self, parent):
        k = 1
        while self.ws.cell(row = k, column = self.column_no+1).value is not None:
            k+=1
        self.row_no = k
        self.current_photo = k-2
    def insert_list_into_excel(self, parent):
        "put the list of categories into the excel file"
        k = 0
        while k < len (self.list):
            self.ws.cell(row=k+2, column=1, value=self.list[k]) #add the saved list of categories to
            k+=1
    def layout (self, parent):
        "Create the basic layout with sizers, buttons, and text boxes"
        self.main_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.main_left = wx.BoxSizer (wx.VERTICAL)
        self.main_right = wx.BoxSizer (wx.VERTICAL)
        sizer_one = wx.BoxSizer (wx.HORIZONTAL)
        sizer_two = wx.BoxSizer (wx.HORIZONTAL)
        prev_btn = wx.Button (self, label = "Prev")
        next_btn = wx.Button (self, label = "Next")
        exit_btn = wx.Button (self, label = "Exit")
        exit_btn.Bind (wx.EVT_BUTTON, parent.dstroy)
        clear_btn = wx.Button (self, label = "Clear")
        clear_btn.Bind (wx.EVT_BUTTON, self.clear_func)
        img = wx.Image (self.max_size, self.max_size)
        self.image_ctrl = wx.StaticBitmap (self, wx.ID_ANY, wx.Bitmap (img))
        self.show_photo_name = wx.TextCtrl (self, size=(200, -1))
        self.main_right.Add (self.show_photo_name,0,wx.ALIGN_CENTER, 15)
        self.main_right.Add(self.image_ctrl, 0, wx.ALL|wx.CENTER, 5)
        self.txtctrl = wx.TextCtrl (self, size=(200, -1))
        next_btn.Bind (wx.EVT_BUTTON, self.next_photo)
        prev_btn.Bind (wx.EVT_BUTTON, self.prev_photo)
        self.txtctrl.Bind (wx.EVT_KEY_DOWN, self.edit_excel)
        self.txtctrl.SetValue (self.current_text)
        sizer_one.Add (prev_btn, 1, wx.ALIGN_CENTER|wx.ALL, 15)
        sizer_one.Add (next_btn, 1, wx.ALIGN_CENTER|wx.ALL, 15)
        sizer_two.Add (self.txtctrl, 1, wx.ALIGN_CENTER|wx.ALL, 15)
        sizer_two.Add (clear_btn, 1, wx.ALIGN_CENTER|wx.ALL, 15)
        for sizer in [sizer_one, sizer_two]:
            self.main_right.Add(sizer, 0, wx.ALIGN_CENTER, 15)
        self.main_right.Add (exit_btn, 0, wx.ALIGN_CENTER, 15)
        self.main_sizer.Add (self.main_left, 1, wx.ALIGN_CENTER|wx.ALIGN_LEFT, 15)
        self.main_sizer.Add (self.main_right, 6.2, wx.ALIGN_CENTER, 15) #
        self.SetSizer (self.main_sizer)
        self.txtctrl.SetFocus()
    def clear_func (self, event):
        self.txtctrl.SetValue ("Type Here")
    def update_photo (self, image):
        "update the image in the bitmap"
        print (self.photos[self.current_photo].split("/")[-1])
        self.show_photo_name.SetValue(self.photos[self.current_photo].split("/")[-1])
        img = wx.Image (image)
        W = img.GetWidth()
        H = img.GetHeight ()
        if W >= H:
            new_W = self.max_size
            new_H = self.max_size*H/W
        else:
            new_H = self.max_size
            new_W = self.max_size*W/H
        img = img.Scale(new_W, new_H)
        self.image_ctrl.SetBitmap(wx.Bitmap(img))
        self.Refresh()
    def list_implementation (self, list):
        "Show and implement the previous list of categories in the textboxes"
        k = 0
        while k < len(list):
            element_name = list[k]
            self.create_new_textbox(element_name, size = 9*len(element_name))
            k+=1
    def create_new_textbox(self, value, size):
        "create a new textbox with -value-"
        if "." in value:
            style = wx.ALIGN_RIGHT
        else:
            style = wx.ALIGN_LEFT
        self.text_box = wx.Button(self, label = value)
        self.text_box.Bind(wx.EVT_BUTTON, self.create_button)
        self.main_left.Add(self.text_box, 0, style|wx.ALL, 10)
    def create_button (self, event):
        button = event.GetEventObject()
        category = button.GetLabel()
        number = category.split(":")[0:1][0]
        if self.txtctrl.GetValue() == "Type Here":
            self.current_text = number
            self.txtctrl.SetValue(category)
        else:
            self.current_text =  self.current_text + "," + number
            prev_text = self.txtctrl.GetValue()
            self.txtctrl.SetValue(prev_text + ", " + category)
    def next_photo (self, event):
        "select the next photo/row"
        self.enter_func(event, go_next_frame = 0)
        self.row_no += 1
        if self.current_photo == self.total_photos - 1:
            print ("end of photos")
        else:
            self.current_photo += 1
        self.current_text == "Type Here"
        self.update_photo(self.photos[self.current_photo])
        self.txtctrl.SetFocus()
    def prev_photo (self, event):
        "select the prev photo/row"
        self.row_no -=1
        if self.current_photo == 0:
            self.current_photo = self.total_photos - 1
        else:
            self.current_photo -= 1
        self.update_photo(self.photos[self.current_photo])
        self.txtctrl.SetFocus()
    def edit_excel (self, event):
        "When typing onto the photo, update the textbox and current_text"
        keycode = event.GetKeyCode()
        if chr(keycode) in string.digits or keycode == 46:
            self.have_found = False
            if self.txtctrl.GetValue() == "Type Here" and keycode != 46:
                self.txtctrl.SetValue (chr(keycode))
                self.current_text = chr(keycode)
                self.previous_number = chr(keycode)
            else:
                self.current_text = self.current_text + chr(keycode)
                current_full_text = self.txtctrl.GetValue()
                self.txtctrl.SetValue (current_full_text + chr(keycode))
            last_number = self.current_text.split(",")[-1:][0]
            n = len(last_number)
            if "." in last_number:
                indexx = last_number.index(".")
                while self.have_found is False and self.finding_counter < len (self.list):
                    for disease in self.list:
                        if str(last_number) == str(disease[0:len(last_number)]):
                            self.have_found = True
                            self.txtctrl.SetValue(self.current_text+str(disease[len(last_number)::]))
                            break
                        self.finding_counter +=1
            elif len(last_number) == 2:
                while self.have_found is False and self.finding_counter < len (self.list):
                    for disease in self.list:
                        if str(last_number) == str(disease[0:len(last_number)]):
                            self.have_found = True
                            self.txtctrl.SetValue(self.current_text+str(disease[len(last_number)::]))
                            break
                        self.finding_counter +=1
                    split_commas = self.txtctrl.GetValue().split(",")
                    length_helper = (len(split_commas)-1)
                    b = split_commas[0:length_helper]
                    counter = 0
                    list = ""
                    while counter <length_helper:
                        list = list+str(split_commas[counter])+", "
                        counter +=1
                    self.txtctrl.SetValue(str(list) + str(self.previous_number) + chr(keycode))

            elif n == 1:
                while self.have_found is False and self.finding_counter < len (self.list):
                    for disease in self.list:
                        if str(last_number) == str(disease[0:n]):
                            self.have_found = True
                            current_full_text = self.txtctrl.GetValue()
                            self.txtctrl.SetValue(str(current_full_text)+str(disease[n::]))
                            break
                        self.finding_counter +=1
            self.finding_counter = 0
            self.previous_number = last_number
        elif keycode == 8:
            n = len (self.current_text) - 1
            self.current_text = self.current_text[0:n]
            if len(self.current_text) == 0:
                self.txtctrl.SetValue("Type Here")
            else:
                self.txtctrl.SetValue(self.current_text)
        elif chr(keycode) == ",":
            if self.txtctrl.GetValue() != "Type Here":
                self.current_text += ","
                current_full_text = self.txtctrl.GetValue()
                self.txtctrl.SetValue (current_full_text + ",")
        elif keycode == 46: #a period
            self.current_text += "."
            self.txtctrl.SetValue (self.current_text)
        elif keycode == 13:
            self.enter_func (event)
        else:
            print ("Only Digits")
    def enter_func (self, event, go_next_frame = 1):
        "When enter is pressed, enter the current_text (sorted) into the excel file"
        adding_string = ""
        for number in self.current_text.split(","):
            for elt in self.list:
                if number == elt[0:1] and "." not in elt:
                    adding_string = adding_string + str(number[0]) +" "
                elif number == elt[0:2] and "." not in elt:
                    adding_string = adding_string + str(number[0:2]) + " "
                elif number == elt[0:3]:
                    adding_string = adding_string + str(number[0:3]) + " "
        split_list = adding_string.split(" ")
        for k in range(len(split_list)-1):
            try:
                self.ws.cell(row=self.row_no, column=self.column_no+k+1, value=int(split_list[k])) #add individual numbers to individual columns (as integers)
            except:
                self.ws.cell(row=self.row_no, column=self.column_no+k+1, value=float(split_list[k])) #add individual numbers to individual columns (as float numbers)
        self.wb.save(self.wb_path)
        self.txtctrl.SetValue("Type Here")
        if go_next_frame:
            self.next_photo (event)
class MainFrame(wx.Frame):
    def __init__ (self):
        super().__init__(None, title='ClassifEye', size = (1000, 800))
        self.sizer = wx.BoxSizer()
        self.SetSizer(self.sizer)
        self.panel = Main_Menu(self)
        self.sizer.Add(self.panel, 1, wx.EXPAND)
        self.excel_static_only_once = True
        self.folder_static_only_once = True
        self.user_static_only_once = True
        self.SetBackgroundColour('black')
        self.Show()
    def dstroy(self, event):
        self.Destroy()
    def one_to_two(self, event):
        if self.panel.select_excel_static_text.GetForegroundColour()!=((3,250,0)):
            if self.excel_static_only_once:
                self.excel_check = wx.StaticText (self, -1, label = "Please Select an Excel File")
                self.excel_check.SetForegroundColour((255,0,0))
                self.excel_advice = wx.StaticText (self, -1, label = "Tip: If creating a new excel file, remember to press enter after typing its name")
                self.excel_advice.SetForegroundColour((246,250,0))
                self.excel_advice_two = wx.StaticText (self, -1, label = "Tip 2: Remember to press 'Open' after finding where you would like to save the file")
                self.excel_advice_two.SetForegroundColour((246,250,0))
                self.panel.main_sizer.Add(self.excel_check,0, wx.CENTER|wx.ALL, 50)
                self.panel.main_sizer.Add(self.excel_advice,0, wx.CENTER|wx.ALL, 0)
                self.panel.main_sizer.Add(self.excel_advice_two,0, wx.CENTER|wx.ALL, 10)
                self.Layout()
                self.excel_static_only_once = False
            else:
                pass
        elif self.panel.select_folder_static_text.GetForegroundColour()!=((3,250,0)):
            if self.folder_static_only_once:
                self.folder_check = wx.StaticText (self, -1, label = "Please Select an Image Directory")
                self.folder_check.SetForegroundColour((255,0,0))
                self.panel.main_sizer.Add(self.folder_check,0, wx.CENTER|wx.ALL, 50)
                self.Layout()
                self.folder_static_only_once = False
            else:
                pass
        elif self.panel.select_user_static_text.GetForegroundColour()!=((3,250,0)):
            if self.user_static_only_once:
                self.user_check = wx.StaticText (self, -1, label = "Please Select a User")
                self.user_check.SetForegroundColour((255,0,0))
                self.panel.main_sizer.Add(self.user_check,0, wx.CENTER|wx.ALL, 50)
                self.Layout()
                self.user_static_only_once = False
            else:
                pass
        else:
            self.panel.Hide()
            self.panel_two = Categories_Page (self, self.panel.list)
            self.sizer.Add(self.panel_two, 1, wx.EXPAND)
            self.panel_two.Show()
            self.Layout()
    def two_to_three(self, event):
        self.panel_two.Hide()
        self.panel_two.check()
        self.panel_three = Excel_Editor (self, photos = self.panel.photos, excel = self.panel.wb_path, user = self.panel.current_user, list = self.panel_two.list)
        self.sizer.Add(self.panel_three, 1, wx.EXPAND)
        self.panel_three.Show()
        self.Layout()
if __name__ == '__main__':
    app = wx.App(redirect=False)
    frame = MainFrame()
    app.MainLoop()
